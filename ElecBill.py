import json
import requests
from dingtalkchatbot.chatbot import DingtalkChatbot
import openpyxl
from datetime import datetime,  timedelta
import schedule
import time

def write_to_excel(remaining_amount):
    """å°†å‰©ä½™ç”µé‡å’ŒæŸ¥è¯¢æ—¶é—´å†™å…¥ Excel è¡¨æ ¼"""
    # æ‰“å¼€æˆ–åˆ›å»º Excel æ–‡ä»¶
    wb = openpyxl.load_workbook('electricity_records.xlsx')
    # è·å–é»˜è®¤çš„å·¥ä½œè¡¨
    sheet = wb.active

    # å¦‚æœä¸å­˜åœ¨åä¸ºâ€œç”µè´¹è®°å½•â€çš„å·¥ä½œè¡¨ï¼Œåˆ™åˆ›å»ºä¸€ä¸ª
    if 'ç”µè´¹è®°å½•' not in wb.sheetnames:
        wb.create_sheet(title='ç”µè´¹è®°å½•')
        sheet = wb['ç”µè´¹è®°å½•']
        # å†™å…¥è¡¨å¤´
        sheet.cell(row=1, column=1, value='å‰©ä½™ç”µé‡')
        sheet.cell(row=1, column=2, value='æŸ¥è¯¢æ—¶é—´')

    # è·å–å½“å‰æ—¶é—´
    query_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # æ‰¾åˆ°ç¬¬ä¸€ä¸ªç©ºè¡Œï¼Œå†™å…¥æ•°æ®
    next_row = sheet.max_row + 1
    sheet.cell(row=next_row, column=1, value=remaining_amount)
    sheet.cell(row=next_row, column=2, value=query_time)

    # ä¿å­˜ Excel æ–‡ä»¶
    wb.save('electricity_records.xlsx')

def get_yesterday_electricity_usage(remaining_amount):
    """è·å–æ˜¨æ—¥ä½¿ç”¨ç”µé‡å¹¶è®¡ç®—æ˜¨æ—¥ç”µè´¹"""
    try:
        # æ‰“å¼€ Excel æ–‡ä»¶
        wb = openpyxl.load_workbook('electricity_records.xlsx')
        sheet = wb['ç”µè´¹è®°å½•']

        yesterday_records = []

        # ä»æœ€åä¸€æ¡å¾€ä¸Šéå†æ¯ä¸€æ¡æ•°æ®
        for row in range(sheet.max_row, 1, -1):
            # è·å–æ—¶é—´å’Œç”µé‡
            record_time = sheet.cell(row=row, column=2).value
            record_amount = sheet.cell(row=row, column=1).value

            # å¦‚æœ record_time æ˜¯å­—ç¬¦ä¸²ç±»å‹ï¼Œè½¬æ¢ä¸º datetime ç±»å‹
            if isinstance(record_time, str):
                record_time = datetime.strptime(record_time, "%Y-%m-%d %H:%M:%S")

            # å¦‚æœè®°å½•çš„æ—¥æœŸæ˜¯æ˜¨å¤©ï¼Œåˆ™å°†å…¶æ·»åŠ åˆ°åˆ—è¡¨ä¸­
            if record_time.date() == datetime.now().date() - timedelta(days=1):
                yesterday_records.append(record_amount)

            # å¦‚æœæ‰¾åˆ°äº†æ˜¨å¤©çš„ç¬¬ä¸€æ¡è®°å½•ï¼ˆå‘ç°æ—¥æœŸæ—©äºå‰å¤©ï¼‰ï¼Œåˆ™è®¡ç®—æ˜¨æ—¥ä½¿ç”¨ç”µé‡å¹¶é€€å‡ºå¾ªç¯
            if record_time.date() < datetime.now().date() - timedelta(days=1):
                if len(yesterday_records) > 1: #å¹¶ä¸”å­˜åœ¨æ˜¨å¤©çš„ä¸¤æ¡è®°å½•
                    yesterday_usage = yesterday_records[-1] - yesterday_records[0]
                    return yesterday_usage
                else:
                    # å¦‚æœæœªæ‰¾åˆ°ç¬¦åˆæ¡ä»¶çš„è®°å½•ï¼Œåˆ™è¿”å›æœªæ‰¾åˆ°æ˜¨æ—¥ç”µè´¹æ•°æ®
                    return "æœªæ‰¾åˆ°æ˜¨æ—¥ç”µè´¹æ•°æ®"

        else:
            # å¦‚æœæœªæ‰¾åˆ°ç¬¦åˆæ¡ä»¶çš„è®°å½•ï¼Œåˆ™è¿”å›æœªæ‰¾åˆ°æ˜¨æ—¥ç”µè´¹æ•°æ®
            return "æœªæ‰¾åˆ°æ˜¨æ—¥ç”µè´¹æ•°æ®"

    except FileNotFoundError:
        return "æœªæ‰¾åˆ°ç”µè´¹è®°å½•æ–‡ä»¶"

def get_useful_electricity_usage(remaining_amount):
    """è·å–å½“å‰ç”µé‡ä¸æ˜¨å¤©æœ€åä¸€æ¡ç”µé‡çš„å·®å€¼ï¼Œç”¨ä»¥ä¼°è®¡æœ¬æ—¥ç”¨é‡"""
    try:
        # æ‰“å¼€ Excel æ–‡ä»¶
        wb = openpyxl.load_workbook('electricity_records.xlsx')
        sheet = wb['ç”µè´¹è®°å½•']

        # ä»æœ€åä¸€æ¡å¾€ä¸Šéå†æ¯ä¸€æ¡æ•°æ®
        for row in range(sheet.max_row, 1, -1):
            # è·å–æ—¶é—´å’Œç”µé‡
            record_time = sheet.cell(row=row, column=2).value
            record_amount = sheet.cell(row=row, column=1).value
            yesterday_last_amount = 0

            # å¦‚æœ record_time æ˜¯å­—ç¬¦ä¸²ç±»å‹ï¼Œè½¬æ¢ä¸º datetime ç±»å‹
            if isinstance(record_time, str):
                record_time = datetime.strptime(record_time, "%Y-%m-%d %H:%M:%S")

            # å¦‚æœè®°å½•çš„æ—¥æœŸæ˜¯æ˜¨å¤©ï¼Œåˆ™å°†å…¶æ·»åŠ åˆ°åˆ—è¡¨ä¸­
            if record_time.date() == datetime.now().date() - timedelta(days=1):
                yesterday_last_amount = record_amount

            # å¦‚æœæ‰¾åˆ°äº†æ˜¨å¤©çš„æœ€åä¸€æ¡è®°å½•ï¼Œåˆ™è®¡ç®—å·®å€¼ç”µé‡å¹¶é€€å‡ºå¾ªç¯
            if yesterday_last_amount != remaining_amount and yesterday_last_amount != 0:
                useful_usage = remaining_amount - yesterday_last_amount
                return useful_usage

            #å¦‚æœæ‰¾åˆ°äº†æ˜¨å¤©ä¹‹å‰çš„è®°å½•ï¼Œè¯´æ˜æ²¡æœ‰æ˜¨å¤©çš„è®°å½•é€€å‡ºå¾ªç¯
            if record_time.date() < datetime.now().date():
                # å¦‚æœæœªæ‰¾åˆ°ç¬¦åˆæ¡ä»¶çš„è®°å½•ï¼Œåˆ™è¿”å›æœªæ‰¾åˆ°æ˜¨æ—¥ç”µè´¹æ•°æ®
                return "æœªæ‰¾åˆ°æ˜¨æ—¥ç”µè´¹æ•°æ®"
        return "æœªæ‰¾åˆ°æ˜¨æ—¥ç”µè´¹æ•°æ®"

    except FileNotFoundError:
        return "æœªæ‰¾åˆ°ç”µè´¹è®°å½•æ–‡ä»¶"

def get_past24hours_electricity_usage(remaining_amount):
    """è·å–è¿‡å»24å°æ—¶å†…ä½¿ç”¨çš„ç”µé‡å¹¶è®¡ç®—æ¶ˆè€—ç”µè´¹"""
    try:
        # æ‰“å¼€ Excel æ–‡ä»¶
        wb = openpyxl.load_workbook('electricity_records.xlsx')
        sheet = wb['ç”µè´¹è®°å½•']

        # éå†æ¯ä¸€æ¡æ•°æ®
        for row in range(sheet.max_row, 1, -1):
            # è·å–æ—¶é—´å’Œç”µé‡
            record_time = sheet.cell(row=row, column=2).value
            record_amount = sheet.cell(row=row, column=1).value

            # æ£€æŸ¥ record_time æ˜¯å¦ä¸ºå­—ç¬¦ä¸²ç±»å‹ï¼Œå¦‚æœæ˜¯åˆ™è½¬æ¢ä¸º datetime ç±»å‹
            if isinstance(record_time, str):
                record_time = datetime.strptime(record_time, "%Y-%m-%d %H:%M:%S")

            # è®¡ç®—ä¸å½“å‰æ—¶é—´çš„æ—¶é—´é—´éš”
            time_difference = datetime.now() - record_time

            # å¦‚æœæ—¶é—´é—´éš”åœ¨ 24 å°æ—¶åˆ° 48 å°æ—¶ä¹‹é—´ï¼Œåˆ™ä½œä¸ºæ˜¨æ—¥çš„è®°å½•
            if timedelta(days=1) <= time_difference < timedelta(days=2):
                # è®¡ç®—æ˜¨æ—¥ä½¿ç”¨ç”µé‡
                yesterday_usage = record_amount - remaining_amount
                return yesterday_usage

            # å¦‚æœæ—¶é—´é—´éš”å¤§äº48å°æ—¶ï¼Œåœæ­¢éå†
            elif time_difference >= timedelta(days=2):
                break

        # å¦‚æœéå†å®Œæ‰€æœ‰æ•°æ®ä»æœªæ‰¾åˆ°ç¬¦åˆæ¡ä»¶çš„è®°å½•ï¼Œåˆ™è¿”å›æœªæ‰¾åˆ°æ˜¨æ—¥ç”µè´¹æ•°æ®
        return "æœªæ‰¾åˆ°æ˜¨æ—¥ç”µè´¹æ•°æ®"

    except FileNotFoundError:
        return "æœªæ‰¾åˆ°ç”µè´¹è®°å½•æ–‡ä»¶"

def check_ifSomebodyPay(remaining_amount):
    """æ£€æŸ¥æ˜¯å¦æœ‰äººå……é’±"""
    try:
        # æ‰“å¼€ Excel æ–‡ä»¶
        wb = openpyxl.load_workbook('electricity_records.xlsx')
        sheet = wb['ç”µè´¹è®°å½•']

        #è¯»å–ä¸Šä¸€æ¡ç”µè´¹è®°å½•
        record_amount = sheet.cell(row=sheet.max_row - 1, column=1).value
        #æ£€æŸ¥æ—¶é—´
        record_time = sheet.cell(row=sheet.max_row, column=2).value
        print(record_time)
        # å¦‚æœé’±å˜å¤šäº†
        if (float(remaining_amount) > float(record_amount)):
            # è®¡ç®—æ˜¨æ—¥ä½¿ç”¨ç”µé‡
            increased_amount = float(remaining_amount) - float(record_amount)
            return increased_amount
        else:
            return 0

    except FileNotFoundError:
        print ("æœªæ‰¾åˆ°ç”µè´¹è®°å½•æ–‡ä»¶")

def check_ifUsageChange(remaining_amount):
    """æ£€æŸ¥æ˜¯å¦æ•°æ®æ›´æ–°"""
    try:
        # æ‰“å¼€ Excel æ–‡ä»¶
        wb = openpyxl.load_workbook('electricity_records.xlsx')
        sheet = wb['ç”µè´¹è®°å½•']

        #è¯»å–ä¸Šä¸€æ¡ç”µè´¹è®°å½•
        record_amount = sheet.cell(row=sheet.max_row - 1, column=1).value
        #record_time = sheet.cell(row=sheet.max_row - 1, column=2).value

        # å¦‚æœé’±æ²¡å˜
        if (float(remaining_amount) == float(record_amount)):
            return False
        else:
            return True

    except FileNotFoundError:
        print ("æœªæ‰¾åˆ°ç”µè´¹è®°å½•æ–‡ä»¶")

def get_electricity_bill():
    """è·å–ç”µè´¹ä¿¡æ¯"""
    url = "http://172.31.248.26:8988/web/Common/Tsm.html"
    headers = {}
    data_dict = {
        "query_elec_roominfo": {
            "aid": "0030000000007301",
            "account": "158086",
            "room": {
                "roomid": room_id,
                "room": "roomid"
            },
            "floor": {
                "floorid": "",
                "floor": ""
            },
            "area": {
                "area": "",
                "areaname": ""
            },
            "building": {
                "buildingid": "",
                "building": ""
            },
            "extdata": "info1="
        }
    }
    # å°†å­—å…¸è½¬æ¢ä¸º JSON å­—ç¬¦ä¸²
    jsondata = json.dumps(data_dict)

    response = requests.post(url, headers=headers, data={"jsondata": jsondata, "funname": "synjones.onecard.query.elec.roominfo"})

    if response.status_code == 200:
        print("ç”µè´¹ä¿¡æ¯è·å–æˆåŠŸ")
        return response.text
    else:
        print("ç”µè´¹ä¿¡æ¯è·å–å¤±è´¥")
        return None

def parse_electricity_bill(bill):
    """è§£æç”µè´¹ä¿¡æ¯"""
    data = json.loads(bill)
    remaining_amount = data['query_elec_roominfo']['errmsg'].split('å‰©ä½™é‡‘é¢:')[1]
    return float(remaining_amount)

def send_notification(remaining_amount, yesterday_usage, increased_amount, useful_usage):
    """å‘é€ç”µè´¹é€šçŸ¥"""
    #ä»…è¿”å›ä¸¤ä½å°æ•°
    remaining_amount = round(remaining_amount, 2)
    yesterday_usage = round(yesterday_usage, 2)
    yesterday_usage = -yesterday_usage
    increased_amount = round(increased_amount, 2)
    useful_usage = round(useful_usage, 2)

    xiaoding = DingtalkChatbot(webhook, secret=secret)
    text = ""

    if remaining_amount < limit:
        #é’±åˆ°è¾¾é˜ˆå€¼
        text += f"âš ï¸ {room} å®¿èˆç”¨ç”µå³å°†æ¬ è´¹ï¼Œè¯·å°½å¿«å……å€¼"
        xiaoding.send_text(text, is_at_all=True)
    else:
        text += f"ğŸ”‹ã€ç”µè´¹ã€‘{room} \n"
        # é’±å˜å¤šäº†
        if increased_amount > 0:
            text += f"ğŸ’°ï¸æœ‰äººå……ç”µè´¹å•¦ï¼ç”µè´¹ä½™é¢å¢åŠ äº† {increased_amount} å…ƒï¼\n"
        #æ­£å¸¸çš„æŠ¥å‘Šä¿¡æ¯
        text += f"ç›®å‰å‰©ä½™ç”µè´¹ {remaining_amount} å…ƒ,\n"
        text += f"æ˜¨æ—¥ç”µè´¹å˜åŒ– {yesterday_usage} å…ƒ\n"
        text += f"å½“å‰ç”µè´¹è¾ƒæ˜¨æ—¥å˜åŒ– {useful_usage} å…ƒã€‚"
        xiaoding.send_text(text)


def main():
    # """ä¸»å‡½æ•°"""
    bill = get_electricity_bill()
    if bill:
        try:
            remaining_amount = parse_electricity_bill(bill)
        except json.decoder.JSONDecodeError:
            print("JSONDecodeError: è§£æç”µè´¹Jsonæ•°æ®å¤±è´¥ï¼")
        print("å‰©ä½™ç”µè´¹:", remaining_amount)
        #è¯»å–æ˜¯å¦å­˜åœ¨æ˜¨æ—¥ç”µè´¹
        yesterday_usage = get_yesterday_electricity_usage(remaining_amount)
        print("æ˜¨æ—¥ç”µè´¹å˜åŒ–:-", yesterday_usage)
        useful_usage = get_useful_electricity_usage(remaining_amount)
        print("å½“å‰ç”µè´¹è¾ƒæ˜¨æ—¥ç”µè´¹å˜åŒ–:-", useful_usage)
        #è¯»å–æ˜¯å¦æœ‰äººå……é’±
        increased_amount = check_ifSomebodyPay(remaining_amount)
        #å¦‚æœæ•°æ®æ›´æ–°ï¼Œå†å‘é€é€šçŸ¥
        if check_ifUsageChange(remaining_amount):
            # å†™å…¥æœ¬åœ°è¡¨æ ¼
            write_to_excel(remaining_amount)
            send_notification(remaining_amount, yesterday_usage, increased_amount, useful_usage)
        # write_to_excel(remaining_amount)
        # send_notification(remaining_amount, yesterday_usage, increased_amount, hours24_usage)
    else:
        print("æœªæŸ¥è¯¢åˆ°ç”µè´¹æ•°æ®")
    print("ç”µè´¹æ£€æŸ¥ç¨‹åºç»“æŸï¼Œä¸‹ä¸€ä¸ªä»»åŠ¡åœ¨ä¸€å°æ—¶å...")

def hourly_job():
    print("å¼€å§‹æ‰§è¡Œå®šæ—¶ç”µé‡æ£€æŸ¥ä»»åŠ¡...")
    main()
    # æ™šä¸Š11ç‚¹æ–­ç½‘å‰å¼ºåˆ¶æ‰§è¡Œå®šæ—¶ä»»åŠ¡
    current_time = datetime.now()
    if current_time.hour == 23:
        print("å½“å‰æ—¶é—´æ˜¯æ™šä¸Š11ç‚¹ï¼Œå¼ºåˆ¶æ‰§è¡Œå®šæ—¶ä»»åŠ¡...")
        main()


if __name__ == "__main__":
    limit = 20  # æ¬ è´¹é¢„è­¦é˜ˆå€¼
    room = '3S527'  # æˆ¿é—´å·
    room_id = '300352711'  # ç”µè´¹æŸ¥è¯¢å·ç 
    # æœºå™¨äººå‚æ•°ï¼Œwebhook å’Œ secretï¼Œä½¿ç”¨æ—¶åœ¨PCç«¯åˆ›å»ºæœºå™¨äººåå¯ä»¥æŸ¥çœ‹å¹¶æ›¿æ¢æˆè‡ªå·±çš„
    webhook = 'https://oapi.dingtalk.com/robot/send?access_token=2e1a8e3bf5c77c5d3e341b63494239d537c8a18dc76653f1231bd52dec4bcfb9'
    secret = 'SEC8834a56af271fb0246db77347726811e3aa13080b888db6c44204db7ab8c0f93'
    main()

    # æ¯å°æ—¶æ‰§è¡Œä¸€æ¬¡
    schedule.every().hour.do(hourly_job)
    # æ— é™å¾ªç¯ä»¥ä¿æŒç¨‹åºè¿è¡Œ
    while True:
        schedule.run_pending()
        time.sleep(1)

